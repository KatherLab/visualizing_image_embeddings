#!/usr/bin/env python3

__author__ = 'Jeff'
__copyright__ = 'Copyright 2023, Kather Lab'
__license__ = 'MIT'
__version__ = '0.1.0'
__maintainer__ = ['Jeff']
__email__ = 'jiefu.zhu@tu-dresden.de'

from pathlib import Path

import numpy as np
# Import openyxl module
import openpyxl

def construct_label_dict():
    # Define variable to load the wookbook
    wookbook = openpyxl.load_workbook("/mnt/sda1/feature-visualizations/JULIEN-POLE-Hypermutated Solid Tumors 2022/PARIS_full_CLINI.xlsx")

    # Define variable to read the active sheet:
    worksheet = wookbook.active
    # put first row and second row into a dictionary
    dict = {}
    # Iterate the loop to read the cell values
    for row in range(2, worksheet.max_row):
        dict[worksheet.cell(row, 1).value] = worksheet.cell(row, 2).value
    print(dict)
    return dict

plot_dir = Path('out_d12')
plot_dir.mkdir(exist_ok=True)
label = np.array([1,2,3],[4,5,6],[7,8,9])
print(label.T.reshape(-1))